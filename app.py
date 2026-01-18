# app.py
import streamlit as st
from pathlib import Path
from file_manager import save_uploaded_file, group_outputs_by_table
from data_processor import (
    preview_file,
    preview_file_with_styles,
    build_combined_highlight_html,
)
from notebook_runner import run_notebook
import mimetypes
import streamlit.components.v1 as components
import shutil
import os
import pandas as pd


# ---------------------
# Paths / config
# ---------------------
ROOT = Path(__file__).parent.resolve()
INPUT_DIR = ROOT / "input_data"
OUTPUT_DIR = ROOT / "highlighted_outputs"
EXPORT_DIR = ROOT / "export_files"

# NOTE: this is the notebook you asked to run
NOTEBOOK_PATH = ROOT / "notebooks" / "Green_Red_Highlighter_9Jan.ipynb"

INPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------
# Helper: wipe outputs folder (safe)
# ---------------------
def wipe_output_dir(output_dir: Path):
    """
    Remove all files and subdirectories inside output_dir but keep the folder itself.
    Returns number of removed items (files+dirs).
    """
    output_dir = Path(output_dir)
    if not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)
        return 0
    removed = 0
    for child in list(output_dir.iterdir()):
        try:
            if child.is_dir():
                shutil.rmtree(child)
            else:
                child.unlink()
            removed += 1
        except Exception as e:
            # log to Streamlit so user sees issues
            st.warning(f"Could not remove {child}: {e}")
    return removed

def get_excel_columns(xlsx_path: Path):
    """Return a list of column names from the first sheet (ignores Unnamed columns)."""
    try:
        df_head = pd.read_excel(xlsx_path, engine="openpyxl", nrows=0, header=0)
        cols = [
            str(c).strip()
            for c in df_head.columns
            if c is not None and not str(c).strip().lower().startswith("unnamed")
        ]
        return cols
    except Exception:
        # fallback if file has weird header formatting
        try:
            df_head = pd.read_excel(xlsx_path, engine="openpyxl", nrows=0, header=None)
            cols = [
                str(c).strip()
                for c in df_head.columns
                if c is not None and not str(c).strip().lower().startswith("unnamed")
            ]
            return cols
        except Exception:
            return []

# ---------------------
# Streamlit UI setup
# ---------------------
st.set_page_config(page_title="Green / Red Highlighter", layout="wide")
st.title("Upload Excel → Run Highlighter Notebook → Download Results")
st.caption(
    "Upload an Excel (.xlsx/.xls). After upload you'll be asked to confirm running the highlighter notebook. "
    "Notebook outputs are written into `highlighted_outputs/` and shown below."
)

# ---------------------
# Project controls
# ---------------------
st.markdown("### Project controls")
st.markdown(
    "Use these buttons to clear generated files.\n\n"
    "- **highlighted_outputs/** → notebook outputs\n"
    "- **export_files/** → final filtered Excel files"
)

col1, col2 = st.columns(2)

with col1:
    if st.button("🟥 Clear highlighted outputs", key="btn_clear_highlighted"):
        n_removed = wipe_output_dir(OUTPUT_DIR)
        if n_removed == 0:
            st.info("Nothing to clear — `highlighted_outputs/` is already empty.")
        else:
            st.success(f"Cleared {n_removed} items from `highlighted_outputs/`.")

with col2:
    if st.button("🟦 Clear export files", key="btn_clear_export"):
        n_removed = wipe_output_dir(EXPORT_DIR)
        if n_removed == 0:
            st.info("Nothing to clear — `export_files/` is already empty.")
        else:
            st.success(f"Cleared {n_removed} items from `export_files/`.")

st.markdown("---")


# ---------------------
# Upload area
# ---------------------
uploaded = st.file_uploader("Upload an Excel file (.xlsx or .xls)", type=["xlsx", "xls"])
# We'll persist the saved path in session_state so reruns keep it
if "uploaded_saved_path" not in st.session_state:
    st.session_state["uploaded_saved_path"] = None

if uploaded is not None:
    # Save upload (fixed filename so notebook can read the same path)
    saved_path = save_uploaded_file(uploaded, INPUT_DIR, filename="uploaded_file.xlsx")
    st.session_state["uploaded_saved_path"] = str(saved_path)
    st.success(f"Saved uploaded file as `{saved_path.name}` in `{INPUT_DIR}`")

# Local variable for convenience
uploaded_saved_path = st.session_state.get("uploaded_saved_path")

st.markdown("### Run the highlighter notebook?")
st.info(
    "The notebook `Green_Red_Highlighter_9Jan.ipynb` will be executed with your uploaded file. "
    "Outputs (Excel + CSV) will be written to the `highlighted_outputs/` folder."
)

# 👇 DEFINE THE BUTTON FIRST
confirm_run = st.button("✅ Confirm & Run notebook", key="btn_run_notebook")

if confirm_run:
    # Guard: ensure the user uploaded a file (saved to input_data/uploaded_file.xlsx)
    if not uploaded_saved_path:
        st.error("No uploaded file found. Please upload a file first (use the uploader above).")
        st.stop()

    # show quick debug info so user can see the path used
    st.info(f"Using uploaded file: `{uploaded_saved_path}`")
    try:
        # list contents (helpful for remote debugging)
        st.write("Contents of input_data:", [p.name for p in INPUT_DIR.iterdir()])
    except Exception:
        pass

    progress = st.progress(0.0)
    progress_msg = st.empty()

    def on_progress(frac, msg):
        try:
            frac = float(frac)
        except Exception:
            frac = 0.0
        frac = max(0.0, min(1.0, frac))
        progress.progress(frac)
        # use info for running messages, success/error replaced below
        progress_msg.info(msg)

    with st.status("Running notebook…", expanded=True) as status:
        success, message = run_notebook(
            NOTEBOOK_PATH,
            Path(uploaded_saved_path),
            OUTPUT_DIR,
            on_progress=on_progress
        )

        if success:
            progress.progress(1.0)
            progress_msg.success("Done!")
            status.update(label="Notebook finished ✅", state="complete", expanded=False)
            st.success("Notebook finished successfully.")
            st.text(message)
        else:
            status.update(label="Notebook failed ❌", state="error", expanded=True)
            st.error("Notebook failed.")
            st.text(message)


# ---------------------
# Output files display (always shown)
# ---------------------
st.markdown("---")
st.markdown("## Notebook output files (from `highlighted_outputs/`)")

st.info(
    "- `*_Green_highlighted.xlsx` = significantly higher (green)\n"
    "- `*_highlighted.xlsx` = significantly lower (red)\n"
    "- CSV masks (`*_highlighted_cells.csv`, `*_sigGreen_highlighted_cells.csv`) help build nicer previews"
)

groups = group_outputs_by_table(OUTPUT_DIR)

# ---------------------
# Table selector (searchable dropdown)
# ---------------------
st.markdown("### Select table (type to search)")

# Only include tables that actually have _Highlights files
all_table_titles = []

for table_title, files in groups.items():
    has_highlights = any(
        "_highlights" in f.name.lower() and "sigtesttable" not in f.name.lower()
        for f in files
    )
    if has_highlights and table_title:
        all_table_titles.append(table_title)

all_table_titles = sorted(all_table_titles)

selected_table = st.selectbox(
    "Start typing to filter tables (e.g. 'A1', 'Table 3', 'Gender')",
    options=["All tables"] + all_table_titles,
    index=0,
    key="table_select"
)

# ---------------------
# Global column filter (applies to all files unless overridden)
# ---------------------
st.markdown("## Global column filter (optional)")
st.caption("Pick columns once and apply to all tables by default. You can still override per table below.")

# Find one sample Highlights file to extract possible columns
all_highlight_files = []
for t in groups.values():
    for f in t:
        if "_highlights" in f.name.lower() and "sigtesttable" not in f.name.lower():
            all_highlight_files.append(f)

sample_file = all_highlight_files[0] if all_highlight_files else None
global_cols_options = get_excel_columns(sample_file) if sample_file else []

use_global_filter = st.checkbox("Enable global column filter", value=False, key="use_global_filter")

global_keep_cols = None
if use_global_filter:
    if not global_cols_options:
        st.warning("Couldn't detect columns for global filter (no Highlights file found yet).")
    else:
        global_keep_cols = st.multiselect(
            "Columns to KEEP (default for all tables)",
            options=global_cols_options,
            default=global_cols_options,
            key="global_keep_cols"
        )

st.markdown("---")

if not groups:
    st.write("No output files found in `highlighted_outputs/` yet.")
else:
    # find global CSV fallbacks (if any)
    global_red_csv = None
    global_green_csv = None
    for c in OUTPUT_DIR.glob("*.csv"):
        name = c.name.lower()
        if "siggreen" in name or "sig_green" in name:
            global_green_csv = c
        elif "highlighted_cells" in name or name.startswith("highlighted"):
            # treat as red/highlighted mask
            global_red_csv = global_red_csv or c

    # iterate tables
    # Decide which tables to show based on dropdown selection
    # Decide which tables to show (only those with _Highlights files)
    if selected_table == "All tables":
        table_titles = all_table_titles
    else:
        table_titles = [selected_table]

        
    for table_title in table_titles:
        # Only keep "Highlights" files AND explicitly exclude SigTestTable outputs
        files = [
            f
            for f in groups[table_title]
            if (
                "_highlights" in f.name.lower()
                and "sigtesttable" not in f.name.lower()
            )
        ]

        # If this table group doesn't contain any valid Highlights file, skip it
        if not files:
            continue

        # Show the table title and proceed
        st.subheader(table_title or "Unknown table")


                # For each file in this table, allow user to pick which columns to keep,
        # then save a filtered copy into export_files/
        EXPORT_DIR = ROOT / "export_files"
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)

        for fpath in files:
            st.write(f"**{fpath.name}**")

            # Try to read just the first row to present sample values / column names
            try:
                sample_df = None
                # read header + first data row (if header present)
                try:
                    sample_df = pd.read_excel(fpath, engine="openpyxl", nrows=1, header=0)
                    header_present = True
                except Exception:
                    # fallback: read without header
                    sample_df = pd.read_excel(fpath, engine="openpyxl", nrows=1, header=None)
                    header_present = False

                if sample_df is None or sample_df.empty:
                    st.info("No data available in this file.")
                    cols = []
                else:
                    # if header_present True, column names are in sample_df.columns
                    cols = [
                        str(c)
                        for c in sample_df.columns
                        if c is not None and not str(c).strip().lower().startswith("unnamed")
                    ]

            except Exception as e:
                st.error(f"Could not read {fpath.name}: {e}")
                cols = []

            # Show a collapsible area with the column selector and sample values
            with st.expander(f"Select columns to KEEP for {fpath.name}", expanded=False):
                if not cols:
                    st.write("No columns detected.")
                else:
                    default_cols = cols.copy()
                    file_key_base = f"keep_{fpath.name}"

                    # Decide what the default selection should be:
                    # - if global filter enabled, use that
                    # - else keep all
                    default_selection = default_cols
                    if use_global_filter and global_keep_cols:
                        # only keep columns that exist in this file
                        default_selection = [c for c in global_keep_cols if c in cols]
                        if not default_selection:
                            default_selection = default_cols  # fallback

                    use_global_for_this_file = False
                    if use_global_filter and global_keep_cols:
                        use_global_for_this_file = st.checkbox(
                            "Use global column selection for this file",
                            value=True,
                            key=f"use_global_{fpath.name}"
                        )

                    if use_global_filter and global_keep_cols and use_global_for_this_file:
                        keep_cols = default_selection
                        st.write(f"Using global selection ({len(keep_cols)} columns).")
                    else:
                        keep_cols = st.multiselect(
                            "Choose columns to KEEP (override)",
                            options=cols,
                            default=default_selection,
                            key=file_key_base
                        )
                    # Always round Base row to whole number
                    decimal_places = 0



                    # Button to save filtered file for this particular fpath
                    # -------------------------
                    # Export / filter controls for this file
                    # -------------------------
                    EXPORT_DIR = ROOT / "export_files"
                    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

                    # unique widget keys based on file name
                    key_base = f"keep_{fpath.name}"

                    # inside the expander we already created earlier we used st.multiselect -> keep_cols variable
                    # so here we show Save button (use a separate key)
                    save_key = f"save_{fpath.name}"

                    if st.button(f"Save filtered copy for {fpath.name}", key=save_key):
                        # Ensure we have the selection from session_state (robust)
                
                        keep_cols = keep_cols

                        if not keep_cols:
                            st.warning("No columns selected — nothing to save.")
                        else:
                            try:
                                import shutil
                                from openpyxl import load_workbook
                                import re

                                # Prepare export path and copy original file into export_files (overwrite)
                                # Prepare export filename: replace _Highlights with _Final
                                export_name = fpath.name
                                export_name = re.sub(r'_Highlights\.xlsx$', '_Final.xlsx', export_name, flags=re.IGNORECASE)

                                dest_path = EXPORT_DIR / export_name
                                shutil.copy2(fpath, dest_path)


                                # --- LOAD workbook and WS first (must set ws before any ws operations) ---
                                wb = load_workbook(filename=str(dest_path))
                                ws = wb.active   # operate on active sheet (this modifies the existing sheet, not create a new one)

                                # --- Insert a new top row and write the cleaned title into A1 ---
                                # (This **shifts the header/data down by 1** so header_row becomes 2)
                                ws.insert_rows(1)

                                # Build clean title from filename: take part before _Highlights, remove underscores
                                stem = fpath.stem
                                clean_title = re.sub(r'(_Highlights|_highlighted|_Green_highlighted|_SigTestTable|_SigTestTable_Highlights).*$', '', stem, flags=re.IGNORECASE)
                                clean_title = clean_title.replace("_", " ").strip()

                                # Write title into first cell of the same sheet (A1)
                                ws.cell(row=1, column=1, value=clean_title)

                                # Header row has moved down by 1
                                header_row = 2

                                # --- Read header values from the header_row (1-based) ---
                                header_values = []
                                max_col = ws.max_column if ws.max_column is not None else 0
                                for col_idx in range(1, max_col + 1):
                                    v = ws.cell(row=header_row, column=col_idx).value
                                    header_values.append("" if v is None else str(v).strip())

                                # Normalize user selection to strings
                                normalized_keep = [str(x).strip() for x in keep_cols]

                                # ALWAYS preserve the first column (index 1)
                                if len(header_values) >= 1:
                                    first_col_name = header_values[0]
                                    if first_col_name not in normalized_keep:
                                        normalized_keep.insert(0, first_col_name)
                                        st.info(f"First column '{first_col_name}' is always preserved and has been added to your selection.")

                                # ALWAYS preserve the "Total" column (case-insensitive)
                                # (If the file doesn't have a Total column, nothing happens.)
                                for h in header_values:
                                    if h and str(h).strip().lower() == "total":
                                        total_col_name = str(h).strip()
                                        if total_col_name not in normalized_keep:
                                            normalized_keep.append(total_col_name)
                                            st.info("Column 'Total' is always preserved and has been added to your selection.")

                                # If header row is empty, abort early to avoid destructive edits
                                if not header_values or all(h == "" for h in header_values):
                                    st.warning(f"No header found in `{fpath.name}` at row {header_row}. Aborting filtered save.")
                                else:
                                    # Build list of 1-based indices to delete (skip first column index 1)
                                    cols_to_delete = [
                                        idx for idx, name in enumerate(header_values, start=1)
                                        if idx != 1 and (name not in normalized_keep)
                                    ]

                                    # Delete from right -> left so indices don't shift unexpectedly
                                    for col_idx in sorted(cols_to_delete, reverse=True):
                                        if 1 <= col_idx <= ws.max_column:
                                            ws.delete_cols(col_idx)
                                    # --- Apply decimal place rounding to the row containing 'Base' ---
                                    try:
                                        target_row = None

                                        # Look for the row that contains the value "Base" (case-insensitive)
                                        for r in range(header_row + 1, ws.max_row + 1):
                                            for c in range(1, ws.max_column + 1):
                                                v = ws.cell(row=r, column=c).value
                                                if v is not None and str(v).strip().lower() == "base":
                                                    target_row = r
                                                    break
                                            if target_row:
                                                break

                                        # If found, round numeric values in that row
                                        if target_row:
                                            for c in range(1, ws.max_column + 1):
                                                cell = ws.cell(row=target_row, column=c)
                                                val = cell.value

                                                if isinstance(val, (int, float)):
                                                    cell.value = round(val, int(decimal_places))
                                                else:
                                                    # handle numeric strings like "12.345"
                                                    try:
                                                        num = float(str(val))
                                                        cell.value = round(num, int(decimal_places))
                                                    except Exception:
                                                        pass  # leave non-numeric values unchanged

                                            st.info("Rounded values in 'Base' row to nearest whole number.")


                                    except Exception as _e:
                                        st.warning(f"Could not apply rounding to 'Base' row: {_e}")

                                    # Save workbook back to dest_path (overwrites copy in export_files)
                                    wb.save(filename=str(dest_path))
                                    wb.close()

                                    st.success(f"Saved filtered file to `export_files/{dest_path.name}` (kept {len(normalized_keep)} columns).")

                                    # Provide immediate download for the edited file
                                    try:
                                        with open(dest_path, "rb") as fh:
                                            st.download_button(
                                                label=f"Download FINAL {dest_path.name}",
                                                data=fh.read(),
                                                file_name=dest_path.name,
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key=f"dl_{dest_path.name}"
                                            )
                                    except Exception as e:
                                        st.warning(f"Saved file but could not create download button: {e}")

                            except Exception as exc:
                                st.error(f"Failed to save filtered copy for {fpath.name}: {exc}")

                    # Original download button for the unchanged highlighted file (always visible)
                    try:
                        data_bytes = fpath.read_bytes()
                        mime = mimetypes.guess_type(fpath.name)[0] or "application/octet-stream"
                        st.download_button(
                            label=f"Download original {fpath.name}",
                            data=data_bytes,
                            file_name=fpath.name,
                            mime=mime,
                            key=f"orig_dl_{fpath.name}"
                        )
                    except Exception as e:
                        st.write(f"_Download not available for original: {e}_")

        st.markdown("---")

